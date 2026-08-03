#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
autosection_compute.py - 分层算量 + 回淤计算（Shapely精确计算）

从C++导出的JSON读取原始数据，使用Shapely进行精确多边形布尔运算，
输出DXF（带HATCH填充）和XLSX（多sheet）。

用法: python autosection_compute.py <input.json>
"""

import json
import sys
import os
import re
import datetime
import unicodedata
import math
import numpy as np

try:
    from shapely.geometry import Polygon, LineString, box, Point
    from shapely.ops import unary_union
except ImportError:
    print("[ERROR] 需要安装shapely: pip install shapely")
    sys.exit(1)

try:
    import ezdxf
    from ezdxf.path import from_hatch, group_paths
except ImportError:
    print("[ERROR] 需要安装ezdxf: pip install ezdxf")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("[ERROR] 需要安装pandas和openpyxl: pip install pandas openpyxl")
    sys.exit(1)

# ==================== 配置常量 ====================

# 地层颜色映射 - 包含"粘"和"黏"两种写法
STRATA_COLORS = {
    '1级淤泥': 11, '1级淤泥质土': 12, '2级淤泥': 31, '3级淤泥': 32,
    '3级粘土': 33, '3级黏土': 33, '4级粘土': 41, '4级黏土': 41,
    '4级淤泥': 42, '5级粘土': 51, '5级黏土': 51,
    '6级砂': 61, '6级碎石': 62, '7级砂': 71, '8级砂': 81, '9级碎石': 91,
    # 填土类型
    '1级填土': 13, '2级填土': 34, '3级填土': 35, '4级填土': 43, '5级填土': 52,
}

HIGH_CONTRAST_COLORS = [
    (255, 0, 0), (0, 200, 0), (0, 0, 255), (255, 255, 0), (255, 0, 255), (0, 255, 255),
    (255, 128, 0), (128, 0, 255), (0, 128, 255), (255, 0, 128), (128, 255, 0), (0, 255, 128),
]

_HATCH_LAYER_CACHE = set()

# Fixed workbook order requested for area and engineering outputs.  Materials
# absent from a drawing are simply omitted; any future unknown material is
# appended after this stable sequence.
FIXED_STRATA_ORDER = (
    '1级淤泥', '2级淤泥', '3级淤泥',
    '1级填土', '4级填土',
    '6级砂', '7级砂', '8级砂',
    '6级碎石', '9级碎石',
    '3级黏土', '4级黏土', '5级黏土',
)


def ordered_strata_layers(layers):
    available = list(dict.fromkeys(layers))
    ordered = [layer for layer in FIXED_STRATA_ORDER if layer in available]
    ordered.extend(layer for layer in available if layer not in ordered)
    return ordered


def four_color_for_layer(layer_name):
    """Return the v4.1 display color for a stratum while keeping its name."""
    if '\u6de4\u6ce5' in layer_name or '\u586b\u571f' in layer_name:
        return 8, (128, 128, 128)       # mud / fill: grey
    if '\u9ecf\u571f' in layer_name:
        return 30, (198, 150, 105)      # clay: brown
    if '\u7802' in layer_name or '\u788e\u77f3' in layer_name:
        return 2, (255, 255, 0)         # sand / gravel: yellow
    return 7, None


# ==================== 工具函数 ====================

def LOG(msg):
    print(msg)


def strata_sort_key(layer_name):
    """地层排序键"""
    m = re.match(r'(\d+)', layer_name)
    return int(m.group(1)) if m else 999


def station_sort_key(station_text):
    """桩号排序键"""
    m = re.match(r'K(\d+)\+(\d+)', station_text)
    if m:
        return int(m.group(1)) * 10000 + int(m.group(2))
    m = re.match(r'S(\d+)', station_text)
    if m:
        return 1000000 + int(m.group(1))
    return 9999999


def get_y_at_x(line, x):
    """在线段上获取指定X处的Y值（线性插值）"""
    if line is None:
        return None
    coords = list(line.coords) if hasattr(line, 'coords') else line
    for i in range(len(coords) - 1):
        x1, y1 = coords[i][0], coords[i][1]
        x2, y2 = coords[i + 1][0], coords[i + 1][1]
        if x1 == x2:
            if abs(x - x1) < 0.01:
                return (y1 + y2) / 2
            continue
        if min(x1, x2) - 0.01 <= x <= max(x1, x2) + 0.01:
            t = (x - x1) / (x2 - x1)
            if -0.001 <= t <= 1.001:
                return y1 + t * (y2 - y1)
    return None


def build_design_polygon(excav_lines, sect_x_min, sect_x_max):
    """构建设计区多边形，底边严格沿开挖线的精确下包络。"""
    if not excav_lines:
        return None

    all_points = [p for l in excav_lines for p in l.coords]
    if not all_points:
        return None

    excav_x_min = min(p[0] for p in all_points)
    excav_x_max = max(p[0] for p in all_points)

    design_x_min = max(excav_x_min, sect_x_min)
    design_x_max = min(excav_x_max, sect_x_max)

    if design_x_max <= design_x_min:
        return None

    boundary_line = _piecewise_envelope(
        excav_lines, 'lower', design_x_min, design_x_max
    )
    if boundary_line is None:
        return None

    boundary_points = list(boundary_line.coords)
    if len(boundary_points) < 2:
        return None

    # 设计区是开挖线以上区域。顶边只用于闭合，取到所有局部开挖线
    # 最高点以上，避免截断实际断面。
    top_y = max(point[1] for point in all_points) + 100.0
    polygon_coords = list(boundary_points)
    polygon_coords.append((boundary_points[-1][0], top_y))
    polygon_coords.append((boundary_points[0][0], top_y))
    polygon_coords.append(polygon_coords[0])

    poly = Polygon(polygon_coords)
    return poly if poly.is_valid else poly.buffer(0)


def _segment_records(lines, x_min=None, x_max=None):
    """把折线拆成X单调线段，保留精确直线方程。"""
    records = []
    for line_index, line in enumerate(lines):
        coords = list(line.coords)
        for segment_index in range(len(coords) - 1):
            x1, y1 = coords[segment_index]
            x2, y2 = coords[segment_index + 1]
            if abs(x2 - x1) < 1e-12:
                continue
            if x1 > x2:
                x1, y1, x2, y2 = x2, y2, x1, y1
            slope = (y2 - y1) / (x2 - x1)
            intercept = y1 - slope * x1
            left = max(x1, x_min) if x_min is not None else x1
            right = min(x2, x_max) if x_max is not None else x2
            if right - left <= 1e-10:
                continue
            records.append(
                (left, right, slope, intercept, line_index, segment_index)
            )
    return records


def _piecewise_envelope(lines, envelope_type, x_min=None, x_max=None):
    """按所有折点和真实交点构造精确上/下包络，不做等距采样。"""
    segments = _segment_records(lines, x_min, x_max)
    if not segments:
        return None

    breakpoints = {value for segment in segments for value in segment[:2]}
    for index, first in enumerate(segments):
        for second in segments[index + 1:]:
            left = max(first[0], second[0])
            right = min(first[1], second[1])
            if right - left <= 1e-10:
                continue
            slope_delta = first[2] - second[2]
            if abs(slope_delta) < 1e-12:
                continue
            crossing = (second[3] - first[3]) / slope_delta
            if left + 1e-9 < crossing < right - 1e-9:
                breakpoints.add(crossing)

    xs = sorted(breakpoints)
    parts = []
    current = []
    choose = min if envelope_type == 'lower' else max

    for left, right in zip(xs, xs[1:]):
        if right - left <= 1e-10:
            continue
        middle = (left + right) / 2.0
        active = [
            segment for segment in segments
            if segment[0] - 1e-9 <= middle <= segment[1] + 1e-9
        ]
        if not active:
            if len(current) >= 2:
                parts.append(current)
            current = []
            continue

        selected = choose(
            active, key=lambda segment: segment[2] * middle + segment[3]
        )
        left_point = (left, selected[2] * left + selected[3])
        right_point = (right, selected[2] * right + selected[3])

        if not current:
            current = [left_point, right_point]
        else:
            previous = current[-1]
            if abs(previous[0] - left) > 1e-7:
                parts.append(current)
                current = [left_point, right_point]
            else:
                if abs(previous[1] - left_point[1]) > 1e-7:
                    current[-1] = (
                        left,
                        min(previous[1], left_point[1])
                        if envelope_type == 'lower'
                        else max(previous[1], left_point[1])
                    )
                current.append(right_point)

    if len(current) >= 2:
        parts.append(current)
    if not parts:
        return None

    # 正常断面连续；若输入存在真正断点，选择横向覆盖最完整的部分，
    # 避免用一条虚构斜线跨越断点。
    points = max(parts, key=lambda part: part[-1][0] - part[0][0])
    compact = []
    for point in points:
        if compact and abs(point[0] - compact[-1][0]) < 1e-9 \
                and abs(point[1] - compact[-1][1]) < 1e-9:
            continue
        compact.append(point)
    return LineString(compact) if len(compact) >= 2 else None


def generate_envelope(dmx_line, update_lines, envelope_type):
    """生成精确包络线（upper=上包络，lower=下包络）。"""
    if not update_lines:
        return dmx_line

    all_lines = [dmx_line] + update_lines
    envelope = _piecewise_envelope(all_lines, envelope_type)
    return envelope if envelope is not None else dmx_line


def detect_ruler_scale_per_section(doc, msp, sect_x_min, sect_x_max, sect_y_min, sect_y_max):
    """逐断面检测标尺比例（对应原始engine_cad_v3.py的RulerDetector.detect_scale）"""
    ruler_layers = ['标尺', '0-标尺', 'RULER']
    ruler_candidates = []

    for layer_name in ruler_layers:
        try:
            for e in msp.query(f'*[layer=="{layer_name}"]'):
                if e.dxftype() == 'INSERT':
                    try:
                        insert_x = e.dxf.insert.x
                        insert_y = e.dxf.insert.y
                        if sect_x_min - 100 <= insert_x <= sect_x_max + 100:
                            y_min, y_max = insert_y, insert_y
                            try:
                                block_name = e.dxf.name
                                if block_name in doc.blocks:
                                    for be in doc.blocks[block_name]:
                                        if be.dxftype() in ('TEXT', 'MTEXT'):
                                            try:
                                                world_y = be.dxf.insert.y + insert_y
                                                y_min = min(y_min, world_y)
                                                y_max = max(y_max, world_y)
                                            except Exception:
                                                pass
                            except Exception:
                                pass
                            ruler_candidates.append({
                                'x': insert_x, 'y_min': y_min, 'y_max': y_max, 'entity': e
                            })
                    except Exception:
                        pass
        except Exception:
            pass

    if not ruler_candidates:
        return None

    # 选择最佳标尺：优先Y重叠最大的，其次选X距离最近的
    sect_y_center = (sect_y_min + sect_y_max) / 2
    sect_x_center = (sect_x_min + sect_x_max) / 2

    best_ruler = None
    best_overlap = -1
    for ruler in ruler_candidates:
        if ruler['y_max'] > ruler['y_min']:
            overlap = max(0, min(sect_y_max, ruler['y_max']) - max(sect_y_min, ruler['y_min']))
            overlap_ratio = overlap / (ruler['y_max'] - ruler['y_min'])
        else:
            overlap_ratio = 0
        if overlap_ratio > best_overlap:
            best_overlap = overlap_ratio
            best_ruler = ruler

    if best_ruler is None or best_overlap <= 0:
        best_ruler = min(ruler_candidates, key=lambda r: abs(r['x'] - sect_x_center))

    # 从最佳标尺提取高程点
    elevation_points = []
    if best_ruler.get('entity'):
        insert_y = best_ruler['entity'].dxf.insert.y
        try:
            block_name = best_ruler['entity'].dxf.name
            if block_name in doc.blocks:
                for be in doc.blocks[block_name]:
                    if be.dxftype() in ('TEXT', 'MTEXT'):
                        try:
                            world_y = be.dxf.insert.y + insert_y
                            text = (be.dxf.text if be.dxftype() == 'TEXT' else be.text).strip()
                            elevation_points.append((world_y, float(text)))
                        except Exception:
                            pass
        except Exception:
            pass

    if len(elevation_points) < 2:
        return None

    # 线性回归: y = a * elevation + b
    n = len(elevation_points)
    sum_y = sum(p[0] for p in elevation_points)
    sum_e = sum(p[1] for p in elevation_points)
    sum_ye = sum(p[0] * p[1] for p in elevation_points)
    sum_e2 = sum(p[1] ** 2 for p in elevation_points)
    denom = n * sum_e2 - sum_e ** 2
    if abs(denom) < 0.001:
        return None

    a = (n * sum_ye - sum_y * sum_e) / denom
    b = (sum_y - a * sum_e) / n

    if abs(a - 5.0) > 2.0 or abs(b - 2510) > 500:
        return None

    return lambda elev: a * elev + b


def shapely_to_boundary_pts(poly):
    """将Shapely多边形转换为边界点列表"""
    if poly.geom_type == 'MultiPolygon':
        # 只返回最大面积的多边形的外边界
        largest = max(poly.geoms, key=lambda g: g.area)
        pts = list(largest.exterior.coords)[:-1]
        return pts
    pts = list(poly.exterior.coords)[:-1]  # 去掉闭合点
    return pts


def add_hatch_to_dxf(msp, poly, layer_name, color_idx=7, rgb_color=None):
    """添加HATCH填充到DXF"""
    # Keep narrow but real fragments so the DXF fill exactly matches the
    # calculated geometry instead of silently losing sub-0.01 m² pieces.
    if poly.is_empty or poly.area <= 1e-9:
        return

    # 处理MultiPolygon：为每个子多边形创建HATCH
    if poly.geom_type == 'MultiPolygon':
        for g in poly.geoms:
            add_hatch_to_dxf(msp, g, layer_name, color_idx, rgb_color)
        return

    layer_key = (id(msp.doc), layer_name)
    if layer_key not in _HATCH_LAYER_CACHE:
        if layer_name not in msp.doc.layers:
            msp.doc.layers.new(name=layer_name, dxfattribs={'color': color_idx})
        else:
            msp.doc.layers.get(layer_name).dxf.color = color_idx
        _HATCH_LAYER_CACHE.add(layer_key)

    pts = shapely_to_boundary_pts(poly)
    if len(pts) < 3:
        return

    try:
        hatch = msp.add_hatch(dxfattribs={'layer': layer_name})
        hatch.set_solid_fill(color=color_idx, rgb=rgb_color)
        hatch.paths.add_polyline_path(pts, is_closed=True)
        # 保留多边形内部孔洞，避免显示面积大于计算面积。
        for interior in poly.interiors:
            hole_pts = list(interior.coords)[:-1]
            if len(hole_pts) >= 3:
                hatch.paths.add_polyline_path(hole_pts, is_closed=True)
    except Exception as e:
        LOG(f"[WARN] HATCH创建失败: {e}")


def _clear_previous_result_entities(msp, elev_layer_name, backfill_layer,
                                    extension_layer="超挖线_延长"):
    """清除输出副本里的旧计算图元，避免重复运行后填充叠加。"""
    removable = []
    result_prefix = f"{elev_layer_name}_"
    for entity in msp:
        layer = entity.dxf.get('layer', '')
        entity_type = entity.dxftype()
        if entity_type == 'HATCH' and (
                layer == backfill_layer or layer.startswith(result_prefix)):
            removable.append(entity)
        elif layer == extension_layer and entity_type in (
                'LINE', 'LWPOLYLINE', 'POLYLINE'):
            removable.append(entity)
        elif layer == elev_layer_name and entity_type in (
                'LINE', 'LWPOLYLINE', 'POLYLINE'):
            removable.append(entity)

    for entity in removable:
        msp.delete_entity(entity)
    if removable:
        LOG(f"[INFO] 已清除旧计算图元: {len(removable)}个")


# ==================== 共享辅助函数 ====================

def _load_hatches_by_layer(msp):
    """预加载所有地层HATCH（按图层分组）"""
    all_hatches_by_layer = {}
    for h in msp.query('HATCH'):
        try:
            layer = h.dxf.layer
            if not re.match(r'^\d', layer):
                continue
            if layer not in all_hatches_by_layer:
                all_hatches_by_layer[layer] = []
            # 一个HATCH可以包含多个外边界、孔洞或曲线边界。先由ezdxf
            # 转成真实路径，再按外边界/孔洞分组，不能只读取第一个路径。
            paths = list(from_hatch(h))
            for path_group in group_paths(paths):
                if not path_group:
                    continue
                outer_pts = [
                    (vertex.x, vertex.y)
                    for vertex in path_group[0].flattening(0.001)
                ]
                if len(outer_pts) < 3:
                    continue
                holes = []
                for hole_path in path_group[1:]:
                    hole_pts = [
                        (vertex.x, vertex.y)
                        for vertex in hole_path.flattening(0.001)
                    ]
                    if len(hole_pts) >= 3:
                        holes.append(hole_pts)
                h_poly = Polygon(outer_pts, holes)
                if not h_poly.is_valid:
                    h_poly = h_poly.buffer(0)
                if not h_poly.is_empty:
                    all_hatches_by_layer[layer].append(h_poly)
        except Exception:
            pass
    return all_hatches_by_layer


def _detect_material_strata_layers(hatches_by_layer):
    """Detect true source strata directly from ezdxf-decoded HATCH layers.

    The Qt DXF reader can lose Chinese layer characters on ANSI_936 drawings;
    using the original HATCH layer names here prevents title-block layers from
    becoming empty strata columns in both DXF and Excel output.
    """
    material_layers = []
    for layer_name in hatches_by_layer:
        candidate = layer_name
        if candidate.startswith('Nonem_'):
            candidate = candidate[len('Nonem_'):]
        if not re.match(r'^\d+', candidate):
            continue
        if any(token in candidate for token in (
            '\u6de4\u6ce5', '\u586b\u571f', '\u9ecf\u571f',
            '\u7802', '\u788e\u77f3',
        )):
            material_layers.append(layer_name)
    return sorted(material_layers, key=strata_sort_key)


def _restore_station_names_from_dxf(doc, sections, pile_layer='0-桩号'):
    """Restore K-stations when the Qt ANSI text reader returns no station text.

    This mirrors StationMatcher::matchSectionToStation: sections arrive in the
    platform's Y-sorted order, every station can be used once, and matching is
    nearest distance with the same 500 CAD-unit tolerance.
    """
    station_pattern = re.compile(r'K?\s*(\d+)\s*\+\s*(\d+)', re.IGNORECASE)
    stations = []
    for entity in doc.modelspace():
        if entity.dxftype() != 'TEXT' or entity.dxf.layer != pile_layer:
            continue
        text = str(entity.dxf.text).strip()
        match = station_pattern.search(text)
        if not match:
            continue
        station_text = f"K{int(match.group(1)):02d}+{int(match.group(2)):03d}"
        insert = entity.dxf.insert
        stations.append((station_text, float(insert.x), float(insert.y)))
    stations.sort(key=lambda item: item[2], reverse=True)

    used = set()
    restored = 0
    for section in sections:
        current = str(section.get('station', ''))
        if current and not re.fullmatch(r'S\d+', current):
            continue
        points = section.get('dmx_points', [])
        if len(points) < 2:
            continue
        center = LineString(points).centroid
        best = None
        best_distance = float('inf')
        for station_text, x_coord, y_coord in stations:
            if station_text in used:
                continue
            distance = math.sqrt(
                (x_coord - center.x) ** 2 * 0.5 +
                (y_coord - center.y) ** 2
            )
            if distance < best_distance and distance < 500.0:
                best = station_text
                best_distance = distance
        if best is not None:
            section['station'] = best
            used.add(best)
            restored += 1
    return restored, len(stations)


def _match_strata_layers(strata_layers, dxf_layers):
    """匹配strata_layers到实际DXF图层名（模糊匹配）"""
    actual_layer_map = {}
    for sl in strata_layers:
        if sl in dxf_layers:
            actual_layer_map[sl] = sl
            continue
        best_match = None
        best_score = 0
        for dl in dxf_layers:
            sl_num = re.match(r'^\d+', sl)
            dl_num = re.match(r'^\d+', dl)
            if sl_num and dl_num and sl_num.group() == dl_num.group():
                sl_rest = sl[sl_num.end():]
                dl_rest = dl[dl_num.end():]
                common = set(sl_rest) & set(dl_rest)
                score = len(common) / max(len(sl_rest), len(dl_rest), 1)
                if score > best_score:
                    best_score = score
                    best_match = dl
        if best_match:
            actual_layer_map[sl] = best_match
        else:
            actual_layer_map[sl] = sl
    return actual_layer_map


def _extend_section(final_section, excav_lines_shapely, overexc_lines_shapely):
    """仅在必要时扩展到本断面左右超挖线的最外侧。"""
    sect_coords_tmp = list(final_section.coords)
    if sect_coords_tmp[0][0] > sect_coords_tmp[-1][0]:
        sect_coords_tmp.reverse()
        final_section = LineString(sect_coords_tmp)
    sect_x_min_tmp = min(c[0] for c in sect_coords_tmp)
    sect_x_max_tmp = max(c[0] for c in sect_coords_tmp)
    sect_y_min_tmp = min(c[1] for c in sect_coords_tmp)
    sect_y_max_tmp = max(c[1] for c in sect_coords_tmp)
    section_box = box(
        sect_x_min_tmp - 20, sect_y_min_tmp - 50,
        sect_x_max_tmp + 20, sect_y_max_tmp + 50
    )
    local_overexc = [
        line for line in overexc_lines_shapely if section_box.intersects(line)
    ]
    local_excav = [
        line for line in excav_lines_shapely if section_box.intersects(line)
    ]
    boundary_lines = local_overexc if local_overexc else local_excav
    if boundary_lines:
        boundary_x_min = min(line.bounds[0] for line in boundary_lines)
        boundary_x_max = max(line.bounds[2] for line in boundary_lines)
        extended_coords = list(sect_coords_tmp)
        if sect_x_min_tmp > boundary_x_min:
            x1, y1 = sect_coords_tmp[0]
            x2, y2 = sect_coords_tmp[1]
            if abs(x2 - x1) > 1e-6:
                slope = (y2 - y1) / (x2 - x1)
                y_ext = y1 + slope * (boundary_x_min - x1)
                extended_coords.insert(0, (boundary_x_min, y_ext))
        if sect_x_max_tmp < boundary_x_max:
            x1, y1 = sect_coords_tmp[-2]
            x2, y2 = sect_coords_tmp[-1]
            if abs(x2 - x1) > 1e-6:
                slope = (y2 - y1) / (x2 - x1)
                y_ext = y2 + slope * (boundary_x_max - x2)
                extended_coords.append((boundary_x_max, y_ext))
        if len(extended_coords) > len(sect_coords_tmp):
            final_section = LineString(extended_coords)
    return final_section


def _geometry_intersection_points(geometry):
    """Extract point candidates from any Shapely intersection result."""
    if geometry is None or geometry.is_empty:
        return []
    if geometry.geom_type == 'Point':
        return [(geometry.x, geometry.y)]
    if geometry.geom_type == 'LineString':
        coords = list(geometry.coords)
        return [coords[0], coords[-1]] if coords else []
    if hasattr(geometry, 'geoms'):
        points = []
        for part in geometry.geoms:
            points.extend(_geometry_intersection_points(part))
        return points
    return []


def _ray_hit_update(endpoint, adjacent, update_geometry, ray_length):
    """Extend an endpoint along its terminal segment and return first hit."""
    dx = endpoint[0] - adjacent[0]
    dy = endpoint[1] - adjacent[1]
    norm = (dx * dx + dy * dy) ** 0.5
    if norm <= 1e-9:
        return None
    ux, uy = dx / norm, dy / norm
    ray_end = (endpoint[0] + ux * ray_length, endpoint[1] + uy * ray_length)
    ray = LineString([endpoint, ray_end])
    candidates = _geometry_intersection_points(ray.intersection(update_geometry))
    forward_hits = []
    for point in candidates:
        distance = (point[0] - endpoint[0]) * ux + (point[1] - endpoint[1]) * uy
        lateral = abs((point[0] - endpoint[0]) * uy -
                      (point[1] - endpoint[1]) * ux)
        if distance >= -1e-7 and lateral <= 1e-5:
            forward_hits.append((max(distance, 0.0), point))
    return min(forward_hits, key=lambda item: item[0])[1] if forward_hits else None


def _extended_overexc_bounds(dmx_line, update_lines, overexc_lines):
    """
    Extend the outer left/right over-excavation endpoints along their original
    terminal directions until they meet the input update section.
    Returns ((left_x, right_x), extension_segments) or (None, []).
    """
    if not update_lines or not overexc_lines:
        return None, []

    update_geometry = unary_union(update_lines)
    if update_geometry.is_empty:
        return None, []

    context_lines = [dmx_line] + list(update_lines)
    all_bounds = [line.bounds for line in context_lines]
    x_min = min(bound[0] for bound in all_bounds)
    y_min = min(bound[1] for bound in all_bounds)
    x_max = max(bound[2] for bound in all_bounds)
    y_max = max(bound[3] for bound in all_bounds)
    width = max(x_max - x_min, 1.0)
    height = max(y_max - y_min, 1.0)
    local_box = box(
        x_min - max(30.0, width * 0.25),
        y_min - max(100.0, height * 2.0),
        x_max + max(30.0, width * 0.25),
        y_max + max(100.0, height * 2.0),
    )
    local_lines = [line for line in overexc_lines if local_box.intersects(line)]
    if not local_lines:
        return None, []

    endpoints = []
    for line in local_lines:
        coords = list(line.coords)
        if len(coords) < 2:
            continue
        endpoints.append((coords[0], coords[1]))
        endpoints.append((coords[-1], coords[-2]))
    if len(endpoints) < 2:
        return None, []

    span = max(width, height, 1.0)
    ray_length = max(1000.0, span * 50.0)

    # Existing contacts are valid stop points. For gaps, try every terminal
    # ray and later choose the outermost valid hit on each side. This is more
    # robust than blindly choosing the smallest/largest endpoint because CAD
    # side slopes are often split into several nearly-connected segments.
    center_x = dmx_line.centroid.x
    hit_candidates = []
    existing_contact = unary_union(local_lines).intersection(update_geometry)
    for hit in _geometry_intersection_points(existing_contact):
        hit_candidates.append((hit, None))
    for endpoint, adjacent in endpoints:
        hit = _ray_hit_update(endpoint, adjacent, update_geometry, ray_length)
        if hit is not None:
            hit_candidates.append((hit, endpoint))

    left_candidates = [
        item for item in hit_candidates
        if item[0][0] < center_x - 1e-7
        and (item[1] is None or item[1][0] < center_x)
    ]
    right_candidates = [
        item for item in hit_candidates
        if item[0][0] > center_x + 1e-7
        and (item[1] is None or item[1][0] > center_x)
    ]
    if not left_candidates or not right_candidates:
        return None, []

    left_hit, left_endpoint = min(
        left_candidates, key=lambda item: item[0][0]
    )
    right_hit, right_endpoint = max(
        right_candidates, key=lambda item: item[0][0]
    )
    left_x, right_x = left_hit[0], right_hit[0]
    if right_x - left_x <= 1e-7:
        return None, []

    segments = []
    if (left_endpoint is not None and
            Point(left_endpoint).distance(Point(left_hit)) > 1e-7):
        segments.append((left_endpoint, left_hit))
    if (right_endpoint is not None and
            Point(right_endpoint).distance(Point(right_hit)) > 1e-7):
        segments.append((right_endpoint, right_hit))
    return (left_x, right_x), segments


def _assign_lines_to_sections(lines, sections):
    """Assign each CAD line to the nearest input DMX section center."""
    section_centers = []
    for section in sections:
        points = section.get('dmx_points', [])
        if len(points) >= 2:
            section_centers.append(LineString(points).centroid)
        else:
            section_centers.append(None)
    assigned = [[] for _ in sections]
    valid_indices = [
        index for index, center in enumerate(section_centers)
        if center is not None
    ]
    if not valid_indices:
        return assigned
    for line in lines:
        center = line.centroid
        nearest = min(
            valid_indices,
            key=lambda index: center.distance(section_centers[index])
        )
        assigned[nearest].append(line)
    return assigned


def _load_source_lines_by_layer(doc, layer_name):
    """Read source CAD curves by their exact Unicode layer name.

    The C++ DXF reader can lose Chinese layer names on some ANSI_936 files.
    The Python stage already reads the source DXF reliably for HATCH and
    stations, so use the same source of truth for geometry needed downstream.
    """
    source_lines = []
    for entity in doc.modelspace():
        if entity.dxf.get('layer', '') != layer_name:
            continue
        try:
            entity_type = entity.dxftype()
            if entity_type == 'LWPOLYLINE':
                points = list(entity.get_points('xy'))
            elif entity_type == 'POLYLINE':
                points = [
                    (vertex.dxf.location.x, vertex.dxf.location.y)
                    for vertex in entity.vertices
                ]
            elif entity_type == 'LINE':
                points = [
                    (entity.dxf.start.x, entity.dxf.start.y),
                    (entity.dxf.end.x, entity.dxf.end.y),
                ]
            else:
                continue
            if len(points) >= 2:
                source_lines.append(LineString(points))
        except Exception as error:
            LOG(f"[WARN] 读取图层{layer_name}曲线失败: {error}")
    return source_lines


def _create_section_polygons(final_section):
    """构建断面多边形，返回(total_open_poly, sect_x_min, sect_x_max, sect_y_min, sect_y_max)"""
    sect_coords = list(final_section.coords)
    sect_x_min = min(c[0] for c in sect_coords)
    sect_x_max = max(c[0] for c in sect_coords)
    sect_y_min = min(c[1] for c in sect_coords)
    sect_y_max = max(c[1] for c in sect_coords)
    bottom_y = sect_y_min - 50
    total_open_poly = Polygon(
        sect_coords + [(sect_x_max, bottom_y), (sect_x_min, bottom_y)]
    )
    if not total_open_poly.is_valid:
        total_open_poly = total_open_poly.buffer(0)
    return total_open_poly, sect_x_min, sect_x_max, sect_y_min, sect_y_max


def _zero_result(station, target_elevation, strata_layers, distinguish_design):
    """返回全零的断面结果"""
    result = {'断面名称': station, '分层线高程': target_elevation, '总面积': 0.0}
    for layer in strata_layers:
        if distinguish_design:
            result[f'{layer}_设计'] = 0.0
            result[f'{layer}_超挖'] = 0.0
        else:
            result[layer] = 0.0
    return result


def _polygon_parts(geometry):
    """Return only non-empty polygon parts from any Shapely geometry."""
    if geometry is None or geometry.is_empty:
        return []
    if geometry.geom_type == 'Polygon':
        return [geometry]
    if hasattr(geometry, 'geoms'):
        return [
            part for part in geometry.geoms
            if part.geom_type == 'Polygon' and not part.is_empty
        ]
    return []


def _process_section(station, target_elevation, strata_layers, distinguish_design,
                     actual_layer_map, all_hatches_by_layer, layer_open,
                     sect_x_min, sect_x_max, sect_y_min, sect_y_max,
                     target_line_y, calc_mode, design_polygon, output_msp,
                     elev_layer_name):
    """处理单个断面的分层算量，返回(result_dict, design_polys, over_polys)"""
    boundary_box = box(sect_x_min - 20, sect_y_min - 50,
                       sect_x_max + 20, sect_y_max + 50)
    strata_areas = {}
    total_area = 0.0
    all_design_polys = []
    all_over_polys = []

    for layer in strata_layers:
        design_area = 0.0
        over_area = 0.0
        total_layer_area = 0.0
        design_polys = []
        over_polys = []
        intersections = []

        actual_layer = actual_layer_map.get(layer, layer)
        hatch_polys = all_hatches_by_layer.get(actual_layer, [])

        for h_poly in hatch_polys:
            try:
                if not boundary_box.intersects(h_poly):
                    continue

                inter = h_poly.intersection(layer_open)
                if inter.is_empty:
                    continue
                intersections.append(inter)
            except Exception:
                pass

        # Overlapping HATCH paths on the same stratum describe one physical
        # region. Merge them before measuring to prevent double counting.
        layer_geometry = unary_union(intersections) if intersections else Polygon()
        if distinguish_design and design_polygon is not None and not design_polygon.is_empty:
            design_geometry = layer_geometry.intersection(design_polygon)
            over_geometry = layer_geometry.difference(design_polygon)
            design_polys = _polygon_parts(design_geometry)
            over_polys = _polygon_parts(over_geometry)
            design_area = design_geometry.area
            over_area = over_geometry.area
        else:
            design_polys = _polygon_parts(layer_geometry)
            total_layer_area = layer_geometry.area

        if distinguish_design:
            strata_areas[f'{layer}_设计'] = round(design_area, 2)
            strata_areas[f'{layer}_超挖'] = round(over_area, 2)
            total_area += design_area + over_area
            all_design_polys.extend(design_polys)
            all_over_polys.extend(over_polys)

            color_idx, rgb_color = four_color_for_layer(layer)
            for poly in design_polys:
                add_hatch_to_dxf(output_msp, poly, f"{elev_layer_name}_{layer}_设计", color_idx, rgb_color)
            for poly in over_polys:
                add_hatch_to_dxf(output_msp, poly, f"{elev_layer_name}_{layer}_超挖", color_idx, rgb_color)
        else:
            strata_areas[layer] = round(total_layer_area, 2)
            total_area += total_layer_area
            all_design_polys.extend(design_polys)

            if total_layer_area > 1e-9:
                color_idx, _ = four_color_for_layer(layer)
                for poly in design_polys:
                    add_hatch_to_dxf(output_msp, poly, f"{elev_layer_name}_{layer}", color_idx)

    result = {
        '断面名称': station,
        '分层线高程': target_elevation,
        **strata_areas,
        '总面积': round(total_area, 2)
    }
    return result, all_design_polys, all_over_polys


def _compute_backfill_for_section(dmx_line, update_lines_shapely, target_line_y,
                                  calc_mode, sect_x_min, sect_x_max,
                                  sect_y_min, sect_y_max, output_msp, backfill_layer,
                                  overexc_lines_shapely=None,
                                  extend_overexc_lines=False,
                                  extension_layer="超挖线_延长"):
    """计算单个断面的回淤面积，返回(面积, 延长状态)。"""
    if not update_lines_shapely:
        return 0.0, 'failed' if extend_overexc_lines else 'disabled'

    upper_envelope = generate_envelope(dmx_line, update_lines_shapely, 'upper')
    if not upper_envelope:
        return 0.0, 'failed' if extend_overexc_lines else 'disabled'

    dmx_coords = list(dmx_line.coords)
    envelope_coords = list(upper_envelope.coords)
    common_x_min = max(min(c[0] for c in dmx_coords), min(c[0] for c in envelope_coords))
    common_x_max = min(max(c[0] for c in dmx_coords), max(c[0] for c in envelope_coords))

    extension_status = 'disabled'
    if extend_overexc_lines:
        extension_bounds, extension_segments = _extended_overexc_bounds(
            dmx_line, update_lines_shapely, overexc_lines_shapely or []
        )
        if extension_bounds is None:
            # Extension is an optional clipping aid.  If it cannot form the
            # two intersections, retain the genuine backfill calculation over
            # the original section range instead of discarding it.
            extension_status = 'fallback'
        else:
            common_x_min = max(common_x_min, extension_bounds[0])
            common_x_max = min(common_x_max, extension_bounds[1])
            extension_status = 'applied'
            if extension_layer not in output_msp.doc.layers:
                output_msp.doc.layers.new(
                    name=extension_layer, dxfattribs={'color': 1}
                )
            for start, end in extension_segments:
                output_msp.add_line(
                    start, end,
                    dxfattribs={'layer': extension_layer, 'color': 1}
                )

    if common_x_max <= common_x_min:
        return 0.0, 'failed' if extend_overexc_lines else extension_status

    # 使用两条折线的全部折点构造带状多边形，不再按0.5m采样。
    x_values = {common_x_min, common_x_max}
    for line in (upper_envelope, dmx_line):
        for x, _ in line.coords:
            if common_x_min - 1e-9 <= x <= common_x_max + 1e-9:
                x_values.add(min(max(x, common_x_min), common_x_max))

    upper_points = []
    dmx_points = []
    for x_current in sorted(x_values):
        envelope_y = get_y_at_x(upper_envelope, x_current)
        dmx_y = get_y_at_x(dmx_line, x_current)
        if envelope_y is not None and dmx_y is not None:
            upper_points.append((x_current, envelope_y))
            dmx_points.append((x_current, dmx_y))

    if len(upper_points) < 2:
        return 0.0, extension_status

    polygon_coords = upper_points + list(reversed(dmx_points))

    if len(polygon_coords) < 3:
        return 0.0, extension_status

    backfill_polygon = Polygon(polygon_coords)
    if not backfill_polygon.is_valid:
        backfill_polygon = backfill_polygon.buffer(0)

    if target_line_y is not None and not backfill_polygon.is_empty:
        if calc_mode == 'below':
            clip_poly = box(common_x_min - 10, sect_y_min - 100,
                            common_x_max + 10, target_line_y)
            backfill_polygon = backfill_polygon.intersection(clip_poly)
        else:
            clip_poly = box(common_x_min - 10, target_line_y,
                            common_x_max + 10, sect_y_max + 100)
            backfill_polygon = backfill_polygon.intersection(clip_poly)

    backfill_area = 0.0
    if not backfill_polygon.is_empty:
        backfill_area = backfill_polygon.area

    if backfill_area > 1e-9 and not backfill_polygon.is_empty:
        add_hatch_to_dxf(output_msp, backfill_polygon, backfill_layer, 4, (0, 255, 255))

    return backfill_area, extension_status


# ==================== 核心计算 ====================

def compute_autosection(data):
    """分层算量计算"""
    input_dxf = data['input_dxf']
    output_dxf = data['output_dxf']
    output_xlsx = data.get('output_xlsx', '')
    target_elevation = data.get('target_elevation', None)
    calc_mode = data.get('calc_mode', 'below')
    distinguish_design = data.get('distinguish_design', False)
    merge_section = data.get('merge_section', False)
    strata_layers = data.get('strata_layers', [])
    sections = data.get('sections', [])
    excav_lines_data = data.get('excav_lines', [])
    overexc_lines_data = data.get('overexc_lines', [])

    LOG(f"[INFO] 分层算量: 高程={target_elevation}, 模式={calc_mode}")
    LOG(f"[INFO] 地层数: {len(strata_layers)}, 断面数: {len(sections)}")

    # 读取源DXF
    doc = ezdxf.readfile(input_dxf)
    output_doc = ezdxf.readfile(input_dxf)
    restored_stations, source_station_count = _restore_station_names_from_dxf(
        doc, sections, data.get('pile_layer', '0-桩号')
    )
    if restored_stations:
        LOG(
            f"[INFO] 从原DXF桩号文本恢复断面名称: "
            f"{restored_stations}/{len(sections)}"
        )
    output_msp = output_doc.modelspace()

    # 创建分层线图层
    elev_layer_name = f"分层线_{target_elevation}m" if target_elevation is not None else "分层线_全算量"
    if elev_layer_name not in output_doc.layers:
        output_doc.layers.new(name=elev_layer_name, dxfattribs={'color': 1})

    # 转换开挖线为Shapely对象
    excav_lines_shapely = []
    for line_data in excav_lines_data:
        pts = line_data['points']
        if len(pts) >= 2:
            excav_lines_shapely.append(LineString(pts))

    # 转换超挖线为Shapely对象
    overexc_lines_shapely = []
    for line_data in overexc_lines_data:
        pts = line_data['points']
        if len(pts) >= 2:
            overexc_lines_shapely.append(LineString(pts))

    msp = doc.modelspace()

    # 预加载HATCH + 匹配图层
    all_hatches_by_layer = _load_hatches_by_layer(msp)
    detected_strata_layers = _detect_material_strata_layers(all_hatches_by_layer)
    if detected_strata_layers:
        strata_layers = ordered_strata_layers(detected_strata_layers)
        LOG(f"[INFO] 从DXF HATCH自动识别地层: {len(strata_layers)}个")
    actual_layer_map = _match_strata_layers(strata_layers, list(all_hatches_by_layer.keys()))
    LOG(f"[INFO] 预加载HATCH: {sum(len(v) for v in all_hatches_by_layer.values())}个, {len(all_hatches_by_layer)}个图层")
    LOG(f"[INFO] 图层映射: {actual_layer_map}")

    results = []

    for idx, section in enumerate(sections):
        station = section['station']
        dmx_pts = section['dmx_points']

        if len(dmx_pts) < 2:
            continue

        dmx_line = LineString(dmx_pts)
        sect_x_min = min(p[0] for p in dmx_pts)
        sect_x_max = max(p[0] for p in dmx_pts)
        sect_y_min = min(p[1] for p in dmx_pts)
        sect_y_max = max(p[1] for p in dmx_pts)

        # 标尺计算高程线Y
        target_line_y = None
        if target_elevation is not None:
            elev_to_y = detect_ruler_scale_per_section(doc, msp, sect_x_min, sect_x_max, sect_y_min, sect_y_max)
            if elev_to_y:
                target_line_y = elev_to_y(target_elevation)
            else:
                target_line_y = 5.0 * target_elevation - 27.0

        # 合并辅助断面线
        final_section = dmx_line
        aux_lines_data = section.get('aux_lines', [])
        if merge_section and aux_lines_data:
            aux_shapely = []
            for pts in aux_lines_data:
                if len(pts) >= 2:
                    aux_shapely.append(LineString(pts))
            if aux_shapely:
                final_section = generate_envelope(dmx_line, aux_shapely, 'lower')

        # 扩展断面线
        final_section = _extend_section(final_section, excav_lines_shapely, overexc_lines_shapely)

        # 构建多边形
        total_open_poly, sect_x_min_a, sect_x_max_a, sect_y_min_a, sect_y_max_a = \
            _create_section_polygons(final_section)

        if total_open_poly.is_empty:
            results.append(_zero_result(station, target_elevation, strata_layers, distinguish_design))
            continue

        # 高程线裁剪
        layer_open = None
        if target_line_y is None:
            layer_open = total_open_poly
        elif calc_mode == 'below':
            if target_line_y < sect_y_min_a:
                results.append(_zero_result(station, target_elevation, strata_layers, distinguish_design))
                if target_line_y > sect_y_min_a - 50:
                    output_msp.add_lwpolyline(
                        [(sect_x_min_a - 5, target_line_y), (sect_x_max_a + 5, target_line_y)],
                        dxfattribs={'layer': elev_layer_name, 'color': 1}
                    )
                continue
            else:
                clip_poly = box(sect_x_min_a - 10, sect_y_min_a - 100,
                                sect_x_max_a + 10, target_line_y)
                layer_open = total_open_poly.intersection(clip_poly)
        else:  # above
            if target_line_y > sect_y_max_a:
                results.append(_zero_result(station, target_elevation, strata_layers, distinguish_design))
                continue
            else:
                clip_poly = box(sect_x_min_a - 10, target_line_y,
                                sect_x_max_a + 10, sect_y_max_a + 100)
                layer_open = total_open_poly.intersection(clip_poly)

        # Debug
        if target_line_y is not None and idx < 5:
            LOG(f"  [{station}] target_y={target_line_y:.1f}, sect_y=[{sect_y_min_a:.1f}, {sect_y_max_a:.1f}], layer_open_area={layer_open.area:.1f}")

        if layer_open.is_empty:
            results.append(_zero_result(station, target_elevation, strata_layers, distinguish_design))
            continue

        # 画高程线
        if target_line_y is not None:
            if calc_mode == 'below' and target_line_y > sect_y_min_a:
                output_msp.add_lwpolyline(
                    [(sect_x_min_a - 5, target_line_y), (sect_x_max_a + 5, target_line_y)],
                    dxfattribs={'layer': elev_layer_name, 'color': 1}
                )
            elif calc_mode == 'above' and target_line_y < sect_y_max_a:
                output_msp.add_lwpolyline(
                    [(sect_x_min_a - 5, target_line_y), (sect_x_max_a + 5, target_line_y)],
                    dxfattribs={'layer': elev_layer_name, 'color': 1}
                )

        # 构建设计区多边形
        design_polygon = None
        if distinguish_design:
            boundary_box = box(sect_x_min_a - 20, sect_y_min_a - 50,
                               sect_x_max_a + 20, sect_y_max_a + 50)
            excav_in_section = [l for l in excav_lines_shapely if boundary_box.intersects(l)]
            if excav_in_section:
                design_polygon = build_design_polygon(excav_in_section, sect_x_min_a, sect_x_max_a)

        # 统计各地层面积
        result, _, _ = _process_section(
            station, target_elevation, strata_layers, distinguish_design,
            actual_layer_map, all_hatches_by_layer, layer_open,
            sect_x_min_a, sect_x_max_a, sect_y_min_a, sect_y_max_a,
            target_line_y, calc_mode, design_polygon, output_msp, elev_layer_name
        )
        results.append(result)

        if (idx + 1) % 50 == 0:
            LOG(f"  已处理 {idx + 1}/{len(sections)} 个断面...")

    # 调试统计
    nonzero_count = sum(1 for r in results if r.get('总面积', 0) > 0)
    total_all = sum(r.get('总面积', 0) for r in results)
    LOG(f"[INFO] 非零断面: {nonzero_count}/{len(results)}, 总面积: {total_all:.1f}")

    # 排序结果
    results.sort(key=lambda x: station_sort_key(x['断面名称']))

    # 保存DXF
    LOG(f"[INFO] 保存DXF: {output_dxf}")
    output_doc.saveas(output_dxf)
    LOG(f"[OK] DXF已保存")

    # 生成Excel
    if results and output_xlsx:
        _write_autosection_excel(results, strata_layers, distinguish_design,
                                 target_elevation, calc_mode, output_xlsx)

    return results


def compute_autosection_backfill(data):
    """分层算量 + 回淤计算合并"""
    input_dxf = data['input_dxf']
    output_dxf = data['output_dxf']
    output_xlsx = data.get('output_xlsx', '')
    target_elevation = data.get('target_elevation', None)
    calc_mode = data.get('calc_mode', 'below')
    distinguish_design = data.get('distinguish_design', False)
    merge_section = data.get('merge_section', False)
    extend_overexc_lines = data.get('extend_overexc_lines', False)
    strata_layers = data.get('strata_layers', [])
    sections = data.get('sections', [])
    excav_lines_data = data.get('excav_lines', [])
    overexc_lines_data = data.get('overexc_lines', [])

    LOG(f"[INFO] 分层+回淤合并: 高程={target_elevation}, 模式={calc_mode}")
    LOG(f"[INFO] 延长超挖线限定回淤: {'开启' if extend_overexc_lines else '关闭'}")
    LOG(f"[INFO] 地层数: {len(strata_layers)}, 断面数: {len(sections)}")

    # 读取源DXF
    doc = ezdxf.readfile(input_dxf)
    output_doc = ezdxf.readfile(input_dxf)
    restored_stations, source_station_count = _restore_station_names_from_dxf(
        doc, sections, data.get('pile_layer', '0-桩号')
    )
    if restored_stations:
        LOG(
            f"[INFO] 从原DXF桩号文本恢复断面名称: "
            f"{restored_stations}/{len(sections)}"
        )
    output_msp = output_doc.modelspace()

    # 创建图层
    elev_layer_name = f"分层线_{target_elevation}m" if target_elevation is not None else "分层线_全算量"
    if elev_layer_name not in output_doc.layers:
        output_doc.layers.new(name=elev_layer_name, dxfattribs={'color': 1})
    backfill_layer = "回淤面积填充"
    if backfill_layer not in output_doc.layers:
        output_doc.layers.new(name=backfill_layer, dxfattribs={'color': 4})
    else:
        output_doc.layers.get(backfill_layer).dxf.color = 4
    _clear_previous_result_entities(
        output_msp, elev_layer_name, backfill_layer
    )

    # 转换开挖线
    excav_lines_shapely = []
    for line_data in excav_lines_data:
        pts = line_data['points']
        if len(pts) >= 2:
            excav_lines_shapely.append(LineString(pts))

    # 转换超挖线
    overexc_lines_shapely = []
    for line_data in overexc_lines_data:
        pts = line_data['points']
        if len(pts) >= 2:
            overexc_lines_shapely.append(LineString(pts))
    if not overexc_lines_shapely:
        overexc_lines_shapely = _load_source_lines_by_layer(doc, '超挖线')
        if overexc_lines_shapely:
            LOG(
                f"[INFO] 从原DXF超挖线图层恢复: "
                f"{len(overexc_lines_shapely)}条"
            )
    overexc_lines_by_section = _assign_lines_to_sections(
        overexc_lines_shapely, sections
    )
    excav_lines_by_section = _assign_lines_to_sections(
        excav_lines_shapely, sections
    )

    msp = doc.modelspace()

    # 预加载HATCH + 匹配图层
    all_hatches_by_layer = _load_hatches_by_layer(msp)
    detected_strata_layers = _detect_material_strata_layers(all_hatches_by_layer)
    if detected_strata_layers:
        strata_layers = ordered_strata_layers(detected_strata_layers)
        LOG(f"[INFO] 从DXF HATCH自动识别地层: {len(strata_layers)}个")
    actual_layer_map = _match_strata_layers(strata_layers, list(all_hatches_by_layer.keys()))
    LOG(f"[INFO] 预加载HATCH: {sum(len(v) for v in all_hatches_by_layer.values())}个, {len(all_hatches_by_layer)}个图层")

    section_results = []
    backfill_results = []
    extension_applied_count = 0
    extension_failed_stations = []

    for idx, section in enumerate(sections):
        station = section['station']
        dmx_pts = section['dmx_points']
        update_lines_data = section.get('update_lines', [])
        local_excav_lines = excav_lines_by_section[idx]
        local_overexc_lines = overexc_lines_by_section[idx]

        if len(dmx_pts) < 2:
            continue

        dmx_line = LineString(dmx_pts)
        sect_x_min = min(p[0] for p in dmx_pts)
        sect_x_max = max(p[0] for p in dmx_pts)
        sect_y_min = min(p[1] for p in dmx_pts)
        sect_y_max = max(p[1] for p in dmx_pts)

        # 标尺计算高程线Y
        target_line_y = None
        if target_elevation is not None:
            elev_to_y = detect_ruler_scale_per_section(doc, msp, sect_x_min, sect_x_max, sect_y_min, sect_y_max)
            if elev_to_y:
                target_line_y = elev_to_y(target_elevation)
            else:
                target_line_y = 5.0 * target_elevation - 27.0

        # 转换更新断面线
        update_lines_shapely = []
        for pts in update_lines_data:
            if len(pts) >= 2:
                update_lines_shapely.append(LineString(pts))

        # 合并断面线
        final_section = dmx_line
        if merge_section and update_lines_shapely:
            final_section = generate_envelope(dmx_line, update_lines_shapely, 'lower')

        # 扩展断面线
        final_section = _extend_section(
            final_section, local_excav_lines, local_overexc_lines
        )

        # 构建多边形
        total_open_poly, sect_x_min_a, sect_x_max_a, sect_y_min_a, sect_y_max_a = \
            _create_section_polygons(final_section)

        if total_open_poly.is_empty:
            section_results.append(_zero_result(station, target_elevation, strata_layers, distinguish_design))
            backfill_results.append({'桩号': station, '回淤面积': 0.0})
            continue

        # 高程线裁剪
        layer_open = None
        if target_line_y is None:
            layer_open = total_open_poly
        elif calc_mode == 'below':
            if target_line_y < sect_y_min_a:
                section_results.append(_zero_result(station, target_elevation, strata_layers, distinguish_design))
                backfill_results.append({'桩号': station, '回淤面积': 0.0})
                if target_line_y > sect_y_min_a - 50:
                    output_msp.add_lwpolyline(
                        [(sect_x_min_a - 5, target_line_y), (sect_x_max_a + 5, target_line_y)],
                        dxfattribs={'layer': elev_layer_name, 'color': 1}
                    )
                continue
            else:
                clip_poly = box(sect_x_min_a - 10, sect_y_min_a - 100,
                                sect_x_max_a + 10, target_line_y)
                layer_open = total_open_poly.intersection(clip_poly)
        else:
            if target_line_y > sect_y_max_a:
                section_results.append(_zero_result(station, target_elevation, strata_layers, distinguish_design))
                backfill_results.append({'桩号': station, '回淤面积': 0.0})
                continue
            else:
                clip_poly = box(sect_x_min_a - 10, target_line_y,
                                sect_x_max_a + 10, sect_y_max_a + 100)
                layer_open = total_open_poly.intersection(clip_poly)

        if layer_open.is_empty:
            section_results.append(_zero_result(station, target_elevation, strata_layers, distinguish_design))
            backfill_results.append({'桩号': station, '回淤面积': 0.0})
            continue

        # 画高程线
        if target_line_y is not None:
            if calc_mode == 'below' and target_line_y > sect_y_min_a:
                output_msp.add_lwpolyline(
                    [(sect_x_min_a - 5, target_line_y), (sect_x_max_a + 5, target_line_y)],
                    dxfattribs={'layer': elev_layer_name, 'color': 1}
                )
            elif calc_mode == 'above' and target_line_y < sect_y_max_a:
                output_msp.add_lwpolyline(
                    [(sect_x_min_a - 5, target_line_y), (sect_x_max_a + 5, target_line_y)],
                    dxfattribs={'layer': elev_layer_name, 'color': 1}
                )

        # 构建设计区多边形
        design_polygon = None
        if distinguish_design:
            boundary_box = box(sect_x_min_a - 20, sect_y_min_a - 50,
                               sect_x_max_a + 20, sect_y_max_a + 50)
            excav_in_section = [
                line for line in local_excav_lines
                if boundary_box.intersects(line)
            ]
            if excav_in_section:
                design_polygon = build_design_polygon(excav_in_section, sect_x_min_a, sect_x_max_a)

        # ===== 分层算量 =====
        result, _, _ = _process_section(
            station, target_elevation, strata_layers, distinguish_design,
            actual_layer_map, all_hatches_by_layer, layer_open,
            sect_x_min_a, sect_x_max_a, sect_y_min_a, sect_y_max_a,
            target_line_y, calc_mode, design_polygon, output_msp, elev_layer_name
        )
        section_results.append(result)

        # ===== 回淤计算 =====
        backfill_area, extension_status = _compute_backfill_for_section(
            dmx_line, update_lines_shapely, target_line_y,
            calc_mode, sect_x_min_a, sect_x_max_a,
            sect_y_min_a, sect_y_max_a, output_msp, backfill_layer,
            local_overexc_lines, extend_overexc_lines
        )
        if extension_status == 'applied':
            extension_applied_count += 1
        elif extension_status in ('failed', 'fallback'):
            extension_failed_stations.append(station)
        backfill_record = {
            '桩号': station,
            '回淤面积': round(backfill_area, 2),
        }
        if extend_overexc_lines:
            backfill_record['范围状态'] = (
                '已按延长超挖线限定'
                if extension_status == 'applied'
                else '未形成左右双交点，按原断面范围计算'
            )
        backfill_results.append(backfill_record)

        if (idx + 1) % 50 == 0:
            LOG(f"  已处理 {idx + 1}/{len(sections)} 个断面...")

    if extend_overexc_lines:
        LOG(f"[INFO] 超挖线延长成功: {extension_applied_count}/{len(sections)}个断面")
        if extension_failed_stations:
            preview = ', '.join(extension_failed_stations[:10])
            suffix = ' ...' if len(extension_failed_stations) > 10 else ''
            LOG(f"[WARN] 未形成左右双交点，已按原断面范围计算回淤: {preview}{suffix}")

    # Excel结果按桩号排序，同时保持分层与回淤数据逐断面对应。
    paired_results = sorted(
        zip(section_results, backfill_results),
        key=lambda pair: station_sort_key(pair[0]['断面名称'])
    )
    if paired_results:
        section_results, backfill_results = map(list, zip(*paired_results))

    # 保存DXF
    LOG(f"[INFO] 保存DXF: {output_dxf}")
    output_doc.saveas(output_dxf)
    LOG(f"[OK] DXF已保存")

    # 生成Excel
    if section_results and output_xlsx:
        _write_combined_excel(section_results, backfill_results, strata_layers,
                              distinguish_design, target_elevation, calc_mode, output_xlsx)

    return section_results, backfill_results


# ==================== Excel输出 ====================

def _excel_display_width(value):
    """按中英文显示宽度估算Excel列宽。"""
    text = '' if value is None else str(value)
    return sum(
        2 if unicodedata.east_asian_width(char) in ('W', 'F', 'A') else 1
        for char in text
    )


def _format_excel_workbook(workbook):
    """统一结果表样式并保证中文标题、数值不被截断。"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    alternate_fill = PatternFill('solid', fgColor='EAF2F8')
    thin_gray = Side(style='thin', color='D9E2F3')
    border = Border(
        left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False
        worksheet.row_dimensions[1].height = 26

        headers = {
            cell.column: str(cell.value or '')
            for cell in worksheet[1]
        }
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True
            )
            cell.border = border

        for row_index, row in enumerate(
                worksheet.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(
                    horizontal='center', vertical='center'
                )
                if row_index % 2 == 0:
                    cell.fill = alternate_fill
                if isinstance(cell.value, (int, float)):
                    header = headers.get(cell.column, '')
                    cell.number_format = '0.00'

        for column_index in range(1, worksheet.max_column + 1):
            values = [
                worksheet.cell(row=row_index, column=column_index).value
                for row_index in range(
                    1, min(worksheet.max_row, 300) + 1
                )
            ]
            width = min(
                26,
                max(10, max((_excel_display_width(v) for v in values), default=8) + 2)
            )
            header = headers.get(column_index, '')
            if '范围状态' in header:
                width = max(width, 24)
            elif column_index == 1:
                width = max(width, 14)
            worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _normalise_local_excel_path(output_xlsx):
    """Return a guaranteed local Windows path for pandas/openpyxl.

    Pandas treats drive paths written as ``D://...`` as fsspec URLs on some
    bundled pandas versions.  The application always writes local files, so
    canonicalise the path and create its parent folder before opening it.
    """
    raw_path = str(output_xlsx).strip().replace('/', os.sep)
    local_path = os.path.abspath(os.path.normpath(raw_path))
    parent_dir = os.path.dirname(local_path)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)
    return local_path


def _write_autosection_excel(results, strata_layers, distinguish_design,
                              target_elevation, calc_mode, output_xlsx):
    """写入分层算量Excel"""
    # QDir can pass a Windows path as D://... .  Pandas interprets that form
    # as an fsspec URL instead of a local drive path, so normalise it first.
    output_xlsx = _normalise_local_excel_path(output_xlsx)
    df = pd.DataFrame(results)
    if distinguish_design:
        component_cols = [
            column for column in df.columns
            if column.endswith('_设计') or column.endswith('_超挖')
        ]
        df['总面积'] = df[component_cols].fillna(0).sum(axis=1).round(2)

    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        if distinguish_design:
            # 设计量sheet
            design_cols = ['断面名称'] + [c for c in df.columns if c.endswith('_设计')]
            df_design = df[design_cols].copy()
            df_design.columns = ['断面名称'] + [c.replace('_设计', '') for c in df.columns if c.endswith('_设计')]
            df_design.to_excel(writer, sheet_name='设计量', index=False)

            # 超挖量sheet
            over_cols = ['断面名称'] + [c for c in df.columns if c.endswith('_超挖')]
            df_over = df[over_cols].copy()
            df_over.columns = ['断面名称'] + [c.replace('_超挖', '') for c in df.columns if c.endswith('_超挖')]
            df_over.to_excel(writer, sheet_name='超挖量', index=False)

            # 总量sheet
            df_total = df[['断面名称']].copy()
            for layer in strata_layers:
                design_col = f'{layer}_设计'
                over_col = f'{layer}_超挖'
                total_val = 0.0
                if design_col in df.columns:
                    total_val = total_val + df[design_col].fillna(0)
                if over_col in df.columns:
                    total_val = total_val + df[over_col].fillna(0)
                df_total[layer] = total_val
            df_total.to_excel(writer, sheet_name='总量', index=False)
        else:
            df.to_excel(writer, sheet_name='明细表', index=False)

        # 地层汇总
        if distinguish_design:
            summary_data = {'地层': [], '设计面积(㎡)': [], '超挖面积(㎡)': []}
            for layer in strata_layers:
                summary_data['地层'].append(layer)
                design_col = f'{layer}_设计'
                over_col = f'{layer}_超挖'
                summary_data['设计面积(㎡)'].append(df[design_col].sum() if design_col in df.columns else 0.0)
                summary_data['超挖面积(㎡)'].append(df[over_col].sum() if over_col in df.columns else 0.0)
            df_summary = pd.DataFrame(summary_data)
            df_summary['总面积(㎡)'] = df_summary['设计面积(㎡)'] + df_summary['超挖面积(㎡)']
        else:
            strata_cols = [c for c in df.columns if '级' in c]
            summary_data = {'地层': strata_cols, '面积(㎡)': [df[c].sum() for c in strata_cols]}
            df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='地层汇总', index=False)

        # 汇总
        mode_text = "以下" if calc_mode == 'below' else "以上"
        total_data = {
            '统计项': ['总断面数', f'{target_elevation}m{mode_text}总面积'],
            '数值': [len(results), df['总面积'].sum()]
        }
        pd.DataFrame(total_data).to_excel(writer, sheet_name='汇总', index=False)
        _format_excel_workbook(writer.book)

    LOG(f"[OK] Excel已保存: {output_xlsx}")

    # 返回总面积供C++解析
    total_area = df['总面积'].sum() if '总面积' in df.columns else 0.0
    print(f"__RESULT__: {json.dumps({'totalArea': round(total_area, 2)})}")
    return total_area


def _ordered_combined_dataframe(dataframe, strata_layers, distinguish_design):
    """Keep all area-related sheets in the requested stable material order."""
    station_columns = [
        column for column in ('断面名称', '桩号', '分层线高程')
        if column in dataframe.columns
    ]
    material_columns = []
    for layer in ordered_strata_layers(strata_layers):
        if distinguish_design:
            material_columns.extend(
                column for column in (f'{layer}_设计', f'{layer}_超挖')
                if column in dataframe.columns
            )
        elif layer in dataframe.columns:
            material_columns.append(layer)
    trailing_columns = [
        column for column in ('回淤面积', '总面积')
        if column in dataframe.columns
    ]
    selected = station_columns + material_columns + trailing_columns
    selected.extend(column for column in dataframe.columns if column not in selected)
    return dataframe.loc[:, selected]


def _write_engineering_sheet(workbook, combined_df):
    """Add the same 25 m trapezoidal engineering-volume formulas as migration."""
    from openpyxl.utils import get_column_letter
    area_coefficient = 0.6

    sheet_name = '\u5de5\u7a0b\u91cf'
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(sheet_name)
    if combined_df.empty:
        worksheet.append(['\u6869\u53f7'])
        return

    station_column = combined_df.columns[0]
    excluded = {
        station_column,
        '\u603b\u9762\u79ef',
        '\u5206\u5c42\u7ebf\u9ad8\u7a0b',
    }
    area_columns = [column for column in combined_df.columns if column not in excluded]
    worksheet.cell(1, 1, '\u6869\u53f7')
    for index, column in enumerate(area_columns):
        base = 2 + index * 2
        worksheet.cell(1, base, f'{column}\u9762\u79ef\u00d70.6(\u33a1)')
        worksheet.cell(1, base + 1, f'{column}\u5de5\u7a0b\u91cf(\u33a5)')

    for row_index, (_, source_row) in enumerate(combined_df.iterrows(), start=2):
        worksheet.cell(row_index, 1, source_row[station_column])
        for index, column in enumerate(area_columns):
            area_col = 2 + index * 2
            volume_col = area_col + 1
            value = source_row[column]
            worksheet.cell(
                row_index, area_col,
                0.0 if pd.isna(value) else round(float(value) * area_coefficient, 2)
            )
            if row_index < len(combined_df) + 1:
                area_letter = get_column_letter(area_col)
                worksheet.cell(
                    row_index, volume_col,
                    f'=ROUND(({area_letter}{row_index}+{area_letter}{row_index + 1})/2*25,2)'
                )

    total_row = len(combined_df) + 2
    worksheet.cell(total_row, 1, '\u5408\u8ba1')
    for index in range(len(area_columns)):
        volume_col = 3 + index * 2
        volume_letter = get_column_letter(volume_col)
        worksheet.cell(total_row, volume_col, f'=SUM({volume_letter}2:{volume_letter}{total_row - 1})')


def _write_combined_excel(section_results, backfill_results, strata_layers,
                           distinguish_design, target_elevation, calc_mode, output_xlsx):
    """写入分层+回淤合并Excel"""
    # See _write_autosection_excel: ensure pandas receives a local path.
    output_xlsx = _normalise_local_excel_path(output_xlsx)
    df_section = pd.DataFrame(section_results)
    df_backfill = pd.DataFrame(backfill_results)
    # Always retain a backfill column and one record per section.  A project
    # may genuinely have zero backfill, but its Excel output must still show
    # explicit 0.00 values instead of dropping the column/sheet data.
    if '回淤面积' not in df_backfill.columns:
        stations = (
            df_section['断面名称'].tolist()
            if '断面名称' in df_section.columns else []
        )
        df_backfill = pd.DataFrame({
            '桩号': stations,
            '回淤面积': [0.0] * len(stations),
        })
    else:
        df_backfill['回淤面积'] = pd.to_numeric(
            df_backfill['回淤面积'], errors='coerce'
        ).fillna(0.0).round(2)
    if distinguish_design:
        component_cols = [
            column for column in df_section.columns
            if column.endswith('_设计') or column.endswith('_超挖')
        ]
        df_section['总面积'] = (
            df_section[component_cols].fillna(0).sum(axis=1).round(2)
        )
        for index, total_area in enumerate(df_section['总面积']):
            section_results[index]['总面积'] = float(total_area)

    df_section = _ordered_combined_dataframe(
        df_section, strata_layers, distinguish_design
    )

    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        # Sheet1: 合并明细表
        combined_results = []
        for i, sr in enumerate(section_results):
            row = dict(sr)
            row['回淤面积'] = (
                backfill_results[i].get('回淤面积', 0.0)
                if i < len(backfill_results) else 0.0
            )
            combined_results.append(row)
        df_combined = pd.DataFrame(combined_results)
        df_combined['回淤面积'] = pd.to_numeric(
            df_combined['回淤面积'], errors='coerce'
        ).fillna(0.0).round(2)
        df_combined = _ordered_combined_dataframe(
            df_combined, strata_layers, distinguish_design
        )
        df_combined.to_excel(writer, sheet_name='合并明细表', index=False)
        _write_engineering_sheet(writer.book, df_combined)

        # Sheet2: 分层算量明细
        if distinguish_design:
            # 设计量sheet
            design_cols = ['断面名称'] + [c for c in df_section.columns if c.endswith('_设计')]
            df_design = df_section[design_cols].copy()
            df_design.columns = ['断面名称'] + [c.replace('_设计', '') for c in df_section.columns if c.endswith('_设计')]
            df_design.to_excel(writer, sheet_name='设计量', index=False)

            # 超挖量sheet
            over_cols = ['断面名称'] + [c for c in df_section.columns if c.endswith('_超挖')]
            df_over = df_section[over_cols].copy()
            df_over.columns = ['断面名称'] + [c.replace('_超挖', '') for c in df_section.columns if c.endswith('_超挖')]
            df_over.to_excel(writer, sheet_name='超挖量', index=False)

            # 总量sheet
            df_total = df_section[['断面名称']].copy()
            for layer in strata_layers:
                design_col = f'{layer}_设计'
                over_col = f'{layer}_超挖'
                total_val = 0.0
                if design_col in df_section.columns:
                    total_val = total_val + df_section[design_col].fillna(0)
                if over_col in df_section.columns:
                    total_val = total_val + df_section[over_col].fillna(0)
                df_total[layer] = total_val
            df_total.to_excel(writer, sheet_name='分层总量', index=False)
        else:
            df_section.to_excel(writer, sheet_name='分层算量明细', index=False)

        # Sheet3: 回淤面积明细
        df_backfill.to_excel(writer, sheet_name='回淤面积明细', index=False)

        # 带合计的回淤
        summary_row = pd.DataFrame([{'桩号': '合计', '回淤面积': df_backfill['回淤面积'].sum()}])
        pd.concat([df_backfill, summary_row], ignore_index=True).to_excel(writer, sheet_name='回淤带合计', index=False)

        # Sheet4: 地层汇总
        if distinguish_design:
            summary_data = {'地层': [], '设计面积(㎡)': [], '超挖面积(㎡)': []}
            for layer in strata_layers:
                summary_data['地层'].append(layer)
                design_col = f'{layer}_设计'
                over_col = f'{layer}_超挖'
                summary_data['设计面积(㎡)'].append(df_section[design_col].sum() if design_col in df_section.columns else 0.0)
                summary_data['超挖面积(㎡)'].append(df_section[over_col].sum() if over_col in df_section.columns else 0.0)
            df_summary = pd.DataFrame(summary_data)
            df_summary['总面积(㎡)'] = df_summary['设计面积(㎡)'] + df_summary['超挖面积(㎡)']
        else:
            strata_cols = [c for c in df_section.columns if '级' in c]
            summary_data = {'地层': strata_cols, '面积(㎡)': [df_section[c].sum() for c in strata_cols]}
            df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='地层汇总', index=False)

        # Sheet5: 总汇总
        mode_text = "以下" if calc_mode == 'below' else "以上"
        total_data = {
            '统计项': [
                '总断面数',
                f'{target_elevation}m{mode_text}总面积' if target_elevation else '开挖总面积',
                '总回淤面积'
            ],
            '数值': [
                len(combined_results),
                df_section['总面积'].sum(),
                df_backfill['回淤面积'].sum()
            ]
        }
        pd.DataFrame(total_data).to_excel(writer, sheet_name='总汇总', index=False)
        _format_excel_workbook(writer.book)

    LOG(f"[OK] Excel已保存: {output_xlsx}")

    # 返回总面积供C++解析
    total_section = df_section['总面积'].sum() if '总面积' in df_section.columns else 0.0
    total_backfill = df_backfill['回淤面积'].sum() if '回淤面积' in df_backfill.columns else 0.0
    print(f"__RESULT__: {json.dumps({'totalArea': round(total_section, 2), 'backfillArea': round(total_backfill, 2)})}")
    return {'section': total_section, 'backfill': total_backfill}


# ==================== 入口 ====================

def main():
    if len(sys.argv) < 2:
        print("Usage: autosection_compute.py <input.json>")
        sys.exit(1)

    json_path = sys.argv[1]

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        task_type = data.get('task_type', 'autosection')

        if task_type == 'autosection':
            compute_autosection(data)
        elif task_type == 'autosection_backfill':
            compute_autosection_backfill(data)
        else:
            print(f"[ERROR] 未知任务类型: {task_type}")
            sys.exit(1)

        sys.exit(0)
    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
