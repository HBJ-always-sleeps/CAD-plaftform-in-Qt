# -*- coding: utf-8 -*-
"""Migrate stratified quantity data into a monthly progress workbook.

The source workbook is the wide-format result produced by the quantity
platform.  The target workbook is used as a template and is never overwritten.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import sys
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import openpyxl


DESIGN_SUFFIX = "_设计"
OVER_SUFFIX = "_超挖"
AREA_QUANT = Decimal("0.01")


def log(message: str) -> None:
    print(message, flush=True)


def normalize_layer(value: Any) -> str:
    text = re.sub(r"\s+", "", str(value or ""))
    return text.replace("黏", "粘").replace("碎石土", "碎石")


def normalize_station(value: Any) -> str:
    text = str(value or "").strip().upper().replace("＋", "+")
    match = re.search(r"K?\s*(\d+)\s*\+\s*(\d+(?:\.\d+)?)", text)
    if not match:
        return re.sub(r"\s+", "", text)

    station = Decimal(match.group(1)) * Decimal(1000) + Decimal(match.group(2))
    normalized = format(station.normalize(), "f")
    return normalized.rstrip("0").rstrip(".") if "." in normalized else normalized


def to_decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal(0)
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(0)


def round_2(value: Decimal | float | int) -> float:
    return float(Decimal(str(value)).quantize(AREA_QUANT, rounding=ROUND_HALF_UP))


def find_source_sheet(workbook: openpyxl.Workbook):
    for preferred in ("合并明细表", "明细表"):
        if preferred in workbook.sheetnames:
            return workbook[preferred]
    for worksheet in workbook.worksheets:
        if "明细" in worksheet.title:
            return worksheet
    return workbook.worksheets[0]


def read_source(source_path: Path) -> tuple[dict[str, dict[str, tuple[Decimal, Decimal]]], str]:
    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    try:
        worksheet = find_source_sheet(workbook)
        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise ValueError("源工作簿为空") from exc

        headers = [str(value).strip() if value is not None else "" for value in raw_headers]
        station_col = next(
            (idx for idx, name in enumerate(headers) if name in {"断面名称", "桩号", "断面桩号"}),
            None,
        )
        if station_col is None:
            station_col = next(
                (idx for idx, name in enumerate(headers) if "桩号" in name or "断面" in name),
                None,
            )
        if station_col is None:
            raise ValueError("源工作簿中未找到断面名称或桩号列")

        layer_columns: list[tuple[str, int, int | None]] = []
        header_index = {name: idx for idx, name in enumerate(headers) if name}
        for idx, name in enumerate(headers):
            if not name.endswith(DESIGN_SUFFIX):
                continue
            layer_name = name[: -len(DESIGN_SUFFIX)]
            over_idx = header_index.get(f"{layer_name}{OVER_SUFFIX}")
            layer_columns.append((normalize_layer(layer_name), idx, over_idx))

        if not layer_columns:
            raise ValueError("源工作簿中未找到“地层_设计”列")

        source_data: dict[str, dict[str, tuple[Decimal, Decimal]]] = {
            layer: {} for layer, _, _ in layer_columns
        }
        for values in rows:
            if station_col >= len(values):
                continue
            station = normalize_station(values[station_col])
            if not station:
                continue
            for layer, design_idx, over_idx in layer_columns:
                design = to_decimal(values[design_idx] if design_idx < len(values) else None)
                over = to_decimal(
                    values[over_idx] if over_idx is not None and over_idx < len(values) else None
                )
                source_data[layer][station] = (design, over)

        return source_data, worksheet.title
    finally:
        workbook.close()


def cached_rows(worksheet, max_row: int, max_col: int = 16) -> list[tuple[Any, ...]]:
    return list(
        worksheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        )
    )


def cached_value(rows: list[tuple[Any, ...]], row: int, col: int) -> Any:
    if row < 1 or col < 1 or row > len(rows) or col > len(rows[row - 1]):
        return None
    return rows[row - 1][col - 1]


def numeric_cached(rows: list[tuple[Any, ...]], row: int, col: int) -> float:
    return round_2(to_decimal(cached_value(rows, row, col)))


def target_station_rows(worksheet) -> list[tuple[int, str]]:
    result: list[tuple[int, str]] = []
    for row in range(1, worksheet.max_row + 1):
        raw_station = worksheet.cell(row=row, column=2).value
        station = normalize_station(raw_station)
        if raw_station is not None and re.fullmatch(r"\d+(?:\.\d+)?", station):
            result.append((row, station))
    return result


def default_output_path(target_path: Path) -> Path:
    timestamp = dt.datetime.now().strftime("%m%d_%H%M")
    return target_path.with_name(f"{target_path.stem}_迁移_{timestamp}{target_path.suffix}")


def migrate(source_path: Path, target_path: Path, coefficient: Decimal, output_path: Path) -> dict[str, Any]:
    if source_path.resolve() == target_path.resolve():
        raise ValueError("源文件和目标文件不能相同")
    if output_path.resolve() == target_path.resolve():
        raise ValueError("输出文件不能覆盖目标模板")

    log(f"读取源文件: {source_path}")
    source_data, source_sheet = read_source(source_path)
    source_station_count = len(next(iter(source_data.values()), {}))
    log(f"源表: {source_sheet}; 地层 {len(source_data)} 个; 桩号 {source_station_count} 个")

    log(f"读取目标模板: {target_path}")
    cached_workbook = openpyxl.load_workbook(target_path, read_only=True, data_only=True)
    workbook = openpyxl.load_workbook(target_path, data_only=False)

    matched_sheets = 0
    matched_stations = 0
    missing_stations = 0
    area_cells = 0
    formula_cells = 0
    unmatched_sheets: list[str] = []
    verification_samples: list[dict[str, Any]] = []

    try:
        cached_by_name = {worksheet.title: worksheet for worksheet in cached_workbook.worksheets}
        for worksheet in workbook.worksheets:
            layer = normalize_layer(worksheet.title)
            layer_data = source_data.get(layer)
            if layer_data is None:
                unmatched_sheets.append(worksheet.title)
                log(f"跳过工作表（源数据无对应地层）: {worksheet.title}")
                continue

            matched_sheets += 1
            station_rows = target_station_rows(worksheet)
            if not station_rows:
                log(f"跳过工作表（未识别到桩号）: {worksheet.title}")
                continue

            cached_sheet = cached_by_name.get(worksheet.title)
            values = cached_rows(cached_sheet, worksheet.max_row) if cached_sheet else []

            # First preserve the previous current-period values in H:K, then
            # write the new source areas into L and N.
            for row, station in station_rows:
                old_l = numeric_cached(values, row, 12)
                old_n = numeric_cached(values, row, 14)
                worksheet.cell(row=row, column=8, value=old_l)
                worksheet.cell(row=row, column=10, value=old_n)

                source_pair = layer_data.get(station)
                if source_pair is None:
                    new_l = 0.0
                    new_n = 0.0
                    missing_stations += 1
                else:
                    new_l = round_2(source_pair[0] * coefficient)
                    new_n = round_2(source_pair[1] * coefficient)
                    matched_stations += 1

                worksheet.cell(row=row, column=12, value=new_l)
                worksheet.cell(row=row, column=14, value=new_n)
                area_cells += 4

            # Each interval formula is anchored on the odd row between two
            # adjacent station-area rows.  This is also the anchor of the
            # vertically merged M/O cells in the supplied template.
            for index in range(len(station_rows) - 1):
                current_row = station_rows[index][0]
                next_row = station_rows[index + 1][0]
                formula_row = current_row + 1

                old_m = numeric_cached(values, formula_row, 13)
                old_o = numeric_cached(values, formula_row, 15)
                if old_m == 0.0:
                    old_m = round_2(
                        (to_decimal(cached_value(values, current_row, 12))
                         + to_decimal(cached_value(values, next_row, 12)))
                        / Decimal(2)
                        * Decimal(25)
                    )
                if old_o == 0.0:
                    old_o = round_2(
                        (to_decimal(cached_value(values, current_row, 14))
                         + to_decimal(cached_value(values, next_row, 14)))
                        / Decimal(2)
                        * Decimal(25)
                    )

                worksheet.cell(row=formula_row, column=9, value=old_m)
                worksheet.cell(row=formula_row, column=11, value=old_o)
                formula_m = f"=ROUND((L{current_row}+L{next_row})/2*25,2)"
                formula_o = f"=ROUND((N{current_row}+N{next_row})/2*25,2)"
                worksheet.cell(row=formula_row, column=13, value=formula_m)
                worksheet.cell(row=formula_row, column=15, value=formula_o)
                formula_cells += 2

                if len(verification_samples) < 5:
                    verification_samples.append(
                        {
                            "sheet": worksheet.title,
                            "cell": f"M{formula_row}",
                            "formula": formula_m,
                        }
                    )

            log(
                f"已迁移: {worksheet.title}; 桩号 {len(station_rows)} 个; "
                f"区间公式 {max(len(station_rows) - 1, 0)} 组"
            )

        if matched_sheets == 0:
            raise ValueError("目标模板中没有与源文件匹配的地层工作表")

        calculation = workbook.calculation
        calculation.calcMode = "auto"
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        cached_workbook.close()
        workbook.close()

    result = {
        "success": True,
        "outputPath": str(output_path),
        "coefficient": float(coefficient),
        "sourceSheet": source_sheet,
        "sourceLayers": len(source_data),
        "sourceStations": source_station_count,
        "matchedSheets": matched_sheets,
        "matchedStations": matched_stations,
        "missingStations": missing_stations,
        "areaCells": area_cells,
        "formulaCells": formula_cells,
        "unmatchedSheets": unmatched_sheets,
        "samples": verification_samples,
    }
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="分层算量结果迁移到月进度工程量表")
    parser.add_argument("source", help="算量结果 Excel")
    parser.add_argument("target", help="月进度工程量表模板")
    parser.add_argument("--coefficient", default="0.6", help="面积系数，默认 0.6")
    parser.add_argument("--output", help="输出路径；省略时在目标文件旁生成新文件")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    source_path = Path(args.source).expanduser()
    target_path = Path(args.target).expanduser()
    output_path = Path(args.output).expanduser() if args.output else default_output_path(target_path)

    try:
        coefficient = Decimal(str(args.coefficient))
        if coefficient <= 0:
            raise ValueError("面积系数必须大于 0")
        if not source_path.is_file():
            raise FileNotFoundError(f"源文件不存在: {source_path}")
        if not target_path.is_file():
            raise FileNotFoundError(f"目标文件不存在: {target_path}")

        result = migrate(source_path, target_path, coefficient, output_path)
        log(f"迁移完成: {result['outputPath']}")
        print("__RESULT__:" + json.dumps(result, ensure_ascii=False), flush=True)
        return 0
    except Exception as exc:  # The Qt caller consumes the structured error.
        result = {"success": False, "error": str(exc)}
        log(f"[ERROR] {exc}")
        print("__RESULT__:" + json.dumps(result, ensure_ascii=False), flush=True)
        return 1


if __name__ == "__main__":
    sys.exit(main())
