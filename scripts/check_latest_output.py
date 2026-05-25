#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""检查最新输出文件"""
import os
import ezdxf
import openpyxl

out_dir = r'D:\断面算量平台\测试文件'

# 找最新文件
files = [(f, os.path.getmtime(os.path.join(out_dir, f))) for f in os.listdir(out_dir) if f.endswith('.xlsx') and '20260522' in f]
files.sort(key=lambda x: x[1], reverse=True)

if files:
    xlsx_path = os.path.join(out_dir, files[0][0])
    print(f"Excel: {files[0][0]}")

    wb = openpyxl.load_workbook(xlsx_path)
    sheets = wb.sheetnames
    print(f"  Sheets: {sheets}")

    # 统计sheet
    ws = wb[sheets[-1]]
    for row in ws.iter_rows(values_only=True):
        print(f"    {row}")

# 找最新DXF
dxf_files = [(f, os.path.getmtime(os.path.join(out_dir, f))) for f in os.listdir(out_dir) if f.endswith('.dxf') and '20260522' in f]
dxf_files.sort(key=lambda x: x[1], reverse=True)

if dxf_files:
    dxf_path = os.path.join(out_dir, dxf_files[0][0])
    print(f"\nDXF: {dxf_files[0][0]}")

    doc = ezdxf.readfile(dxf_path)
    msp = doc.modelspace()

    # 统计图层
    hatch_layers = {}
    for h in msp.query('HATCH'):
        layer = h.dxf.layer
        hatch_layers[layer] = hatch_layers.get(layer, 0) + 1

    design_count = sum(v for k, v in hatch_layers.items() if '设计' in k)
    over_count = sum(v for k, v in hatch_layers.items() if '超挖' in k)

    print(f"  HATCH: 设计={design_count}, 超挖={over_count}")

    # 统计高程线
    for layer in doc.layers:
        name = layer.dxf.name
        if '分层线' in name:
            count = len(list(msp.query(f'LWPOLYLINE[layer=="{name}"]')))
            if count > 0:
                print(f"  分层线 {name}: {count}")