#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""测试颜色匹配"""
import os
import json

out_dir = r'D:\断面算量平台\测试文件'
xlsx_files = [(f, os.path.getmtime(os.path.join(out_dir, f))) for f in os.listdir(out_dir) if f.endswith('.xlsx') and '20260522' in f and '-4m' in f]
xlsx_files.sort(key=lambda x: x[1], reverse=True)

if xlsx_files:
    import openpyxl
    xlsx_path = os.path.join(out_dir, xlsx_files[0][0])
    wb = openpyxl.load_workbook(xlsx_path)

    # 检查所有sheets
    print(f"Sheets: {wb.sheetnames}")

    # 从各sheet获取地层名
    strata_names = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] and isinstance(row[0], str) and row[0].startswith(('1级', '2级', '3级', '4级', '5级', '6级', '7级', '8级', '9级')):
                strata_names.append(row[0])

    print("\n地层名称（从Excel读取）:")
    for name in strata_names[:10]:
        # 打印Unicode码点
        codes = ' '.join([f'U+{ord(c):04X}' for c in name])
        print(f"  '{name}' -> {codes}")

# STRATA_COLORS的定义
STRATA_COLORS = {
    '1级淤泥': 11, '1级淤泥质土': 12, '2级淤泥': 31, '3级淤泥': 32,
    '3级粘土': 33, '4级粘土': 41, '4级淤泥': 42, '5级粘土': 51,
    '6级砂': 61, '6级碎石': 62, '7级砂': 71, '8级砂': 81, '9级碎石': 91,
}

print("\nSTRATA_COLORS键:")
for key in list(STRATA_COLORS.keys())[:5]:
    codes = ' '.join([f'U+{ord(c):04X}' for c in key[:5]])
    print(f"  '{key}' -> {codes}")

# 测试匹配
print("\n匹配测试:")
for name in strata_names[:10]:
    color = STRATA_COLORS.get(name, 7)
    print(f"  '{name}' -> color={color}")