#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
对比Python原版和Qt重构版的输出文件
比较内容：Excel结构、DXF图层、HATCH数量、颜色、实体数
"""
import os
import ezdxf
import openpyxl
from collections import defaultdict

out_dir = r'D:\断面算量平台\测试文件'

def get_latest_file(pattern, date_str, keyword=''):
    """获取指定日期最新的文件"""
    files = [(f, os.path.getmtime(os.path.join(out_dir, f)))
             for f in os.listdir(out_dir)
             if f.endswith(pattern) and date_str in f and (keyword == '' or keyword in f)]
    files.sort(key=lambda x: x[1], reverse=True)
    return files[0][0] if files else None

def compare_xlsx(py_file, qt_file):
    """对比Excel文件"""
    print("=" * 60)
    print("Excel文件对比")
    print("=" * 60)
    print(f"  Python原版: {py_file}")
    print(f"  Qt重构版:   {qt_file}")
    print()

    wb_py = openpyxl.load_workbook(os.path.join(out_dir, py_file))
    wb_qt = openpyxl.load_workbook(os.path.join(out_dir, qt_file))

    # 对比sheets
    py_sheets = wb_py.sheetnames
    qt_sheets = wb_qt.sheetnames
    print(f"  Sheet数量: Python={len(py_sheets)}, Qt={len(qt_sheets)}")

    if py_sheets != qt_sheets:
        print(f"  [差异] Sheet名称不同!")
        print(f"    Python: {py_sheets}")
        print(f"    Qt:     {qt_sheets}")
    else:
        print(f"  [一致] Sheet名称相同: {py_sheets}")

    # 对比每个sheet的内容
    common_sheets = set(py_sheets) & set(qt_sheets)
    for sheet_name in sorted(common_sheets):
        ws_py = wb_py[sheet_name]
        ws_qt = wb_qt[sheet_name]

        py_rows = ws_py.max_row
        qt_rows = ws_qt.max_row
        py_cols = ws_py.max_column
        qt_cols = ws_qt.max_column

        status = "[一致]" if py_rows == qt_rows and py_cols == qt_cols else "[差异]"
        print(f"\n  Sheet: {sheet_name}")
        print(f"    行数: Python={py_rows}, Qt={qt_rows} {status}")
        print(f"    列数: Python={py_cols}, Qt={qt_cols} {status}")

        # 对比前几行数据
        if py_rows == qt_rows and py_cols == qt_cols:
            mismatch_count = 0
            for row in range(1, min(py_rows + 1, 50)):
                for col in range(1, py_cols + 1):
                    val_py = ws_py.cell(row=row, column=col).value
                    val_qt = ws_qt.cell(row=row, column=col).value
                    if val_py != val_qt:
                        if mismatch_count < 5:
                            print(f"    [差异] 行{row}列{col}: Python='{val_py}', Qt='{val_qt}'")
                        mismatch_count += 1
            if mismatch_count == 0:
                print(f"    [一致] 前{min(py_rows, 50)}行数据完全一致")
            else:
                print(f"    [差异] 共{mismatch_count}处数据不同")

def compare_dxf(py_file, qt_file):
    """对比DXF文件"""
    print("\n" + "=" * 60)
    print("DXF文件对比")
    print("=" * 60)
    print(f"  Python原版: {py_file}")
    print(f"  Qt重构版:   {qt_file}")
    print()

    doc_py = ezdxf.readfile(os.path.join(out_dir, py_file))
    doc_qt = ezdxf.readfile(os.path.join(out_dir, qt_file))

    msp_py = doc_py.modelspace()
    msp_qt = doc_qt.modelspace()

    # 统计实体类型
    def count_entities(msp):
        counts = defaultdict(int)
        for e in msp:
            counts[e.dxftype()] += 1
        return dict(counts)

    py_entities = count_entities(msp_py)
    qt_entities = count_entities(msp_qt)

    all_types = sorted(set(list(py_entities.keys()) + list(qt_entities.keys())))
    print("  实体类型统计:")
    for etype in all_types:
        py_count = py_entities.get(etype, 0)
        qt_count = qt_entities.get(etype, 0)
        status = "[一致]" if py_count == qt_count else "[差异]"
        if py_count > 0 or qt_count > 0:
            print(f"    {etype}: Python={py_count}, Qt={qt_count} {status}")

    # 对比图层
    def get_layers(doc):
        layers = {}
        for layer in doc.layers:
            name = layer.dxf.name
            color = layer.color if hasattr(layer, 'color') else layer.dxf.color
            layers[name] = color
        return layers

    py_layers = get_layers(doc_py)
    qt_layers = get_layers(doc_qt)

    all_layer_names = sorted(set(list(py_layers.keys()) + list(qt_layers.keys())))
    print(f"\n  图层数量: Python={len(py_layers)}, Qt={len(qt_layers)}")

    # 只显示有差异的图层
    diff_layers = []
    for name in all_layer_names:
        py_color = py_layers.get(name)
        qt_color = qt_layers.get(name)
        if py_color != qt_color:
            diff_layers.append((name, py_color, qt_color))

    if diff_layers:
        print(f"  [差异] 有{len(diff_layers)}个图层颜色不同:")
        for name, py_c, qt_c in diff_layers[:10]:
            print(f"    {name}: Python={py_c}, Qt={qt_c}")
    else:
        print(f"  [一致] 所有图层颜色一致")

    # 对比HATCH详细信息
    def analyze_hatches(msp):
        hatches = defaultdict(lambda: {'count': 0, 'colors': set(), 'areas': []})
        for h in msp.query('HATCH'):
            layer = h.dxf.layer
            color = h.dxf.color
            hatches[layer]['count'] += 1
            hatches[layer]['colors'].add(color)
        return dict(hatches)

    py_hatches = analyze_hatches(msp_py)
    qt_hatches = analyze_hatches(msp_qt)

    all_hatch_layers = sorted(set(list(py_hatches.keys()) + list(qt_hatches.keys())))
    print(f"\n  HATCH图层统计:")
    for layer in all_hatch_layers:
        py_info = py_hatches.get(layer, {'count': 0, 'colors': set()})
        qt_info = qt_hatches.get(layer, {'count': 0, 'colors': set()})
        py_count = py_info['count']
        qt_count = qt_info['count']
        py_colors = sorted(py_info['colors'])
        qt_colors = sorted(qt_info['colors'])
        count_status = "[一致]" if py_count == qt_count else "[差异]"
        color_status = "[一致]" if py_colors == qt_colors else "[差异]"
        if py_count > 0 or qt_count > 0:
            print(f"    {layer}: 数量 Python={py_count}, Qt={qt_count} {count_status}")
            print(f"      颜色: Python={py_colors}, Qt={qt_colors} {color_status}")

def main():
    print("=" * 60)
    print("Python原版 vs Qt重构版 输出对比")
    print("=" * 60)

    # Test 1: 0m分层回淤
    print("\n" + "#" * 60)
    print("测试1: 0m分层回淤 (全算量)")
    print("#" * 60)

    py_xlsx_1 = get_latest_file('.xlsx', '20260501', '0m')
    qt_xlsx_1 = get_latest_file('.xlsx', '20260522', '0m')
    if py_xlsx_1 and qt_xlsx_1:
        compare_xlsx(py_xlsx_1, qt_xlsx_1)

    py_dxf_1 = get_latest_file('.dxf', '20260501', '0m')
    qt_dxf_1 = get_latest_file('.dxf', '20260522', '0m')
    if py_dxf_1 and qt_dxf_1:
        compare_dxf(py_dxf_1, qt_dxf_1)

    # Test 2: -4m以上面积
    print("\n" + "#" * 60)
    print("测试2: -4m以上面积 (分层算量)")
    print("#" * 60)

    py_xlsx_2 = get_latest_file('.xlsx', '20260501', '-4m')
    qt_xlsx_2 = get_latest_file('.xlsx', '20260522', '-4m')
    if py_xlsx_2 and qt_xlsx_2:
        compare_xlsx(py_xlsx_2, qt_xlsx_2)

    py_dxf_2 = get_latest_file('.dxf', '20260501', '-4m')
    qt_dxf_2 = get_latest_file('.dxf', '20260522', '-4m')
    if py_dxf_2 and qt_dxf_2:
        compare_dxf(py_dxf_2, qt_dxf_2)

if __name__ == '__main__':
    main()
